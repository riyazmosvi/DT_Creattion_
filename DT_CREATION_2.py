import pandas as pd
import streamlit as st
import re
from io import BytesIO
from openpyxl import Workbook

st.set_page_config(page_title="DT Creation Tool", layout="wide")
st.title("🔌 DT Creation Tool")

# Normalize DT name
def normalize(dt):
    if pd.isna(dt): return ""
    dt = str(dt).strip()
    match = re.search(r"(\d+\s*KVA)", dt, re.IGNORECASE)
    if match:
        cap = match.group(1).replace(" ", "").upper()
        rest = re.sub(match.group(1), "", dt, flags=re.IGNORECASE).strip()
        return f"{cap} {rest}".upper()
    return dt.upper()

# Generate DT codes
def increment_dt_codes(start, count):
    match = re.match(r"([A-Za-z\-]+)(\d+)", start.strip().upper())
    if match:
        pre, num = match.groups()
        num = int(num)
        width = len(match.group(2))
        return [f"{pre}{str(num + i).zfill(width)}" for i in range(count)]
    return [f"{start}-{i}" for i in range(count)]

# Upload file
uploaded_file = st.file_uploader("📁 Upload the Excel file", type=[".xlsx"])
if not uploaded_file:
    st.info("Please upload an Excel file to begin.")
    st.stop()

df_sheets = pd.read_excel(uploaded_file, sheet_name=None)
sheet = st.selectbox("📑 Select Sheet", df_sheets.keys())
data_raw = df_sheets[sheet]
st.dataframe(data_raw.head(10), use_container_width=True)

# Column mapping
st.header("🔧 Select Columns")
cols = [c for c in data_raw.columns if not data_raw[c].isna().all()]
dt_col = st.selectbox("DT Name column", cols)
consumer_col = st.selectbox("Consumer Code column", cols)
feeder_col = st.selectbox("Feeder Code column", cols)
mr_name_col = st.selectbox("MR Name column", cols)
mr_code_col = st.selectbox("MR Code column", cols)

# 🟡 Step 1: Forward fill DT_NAME & FEEDER_CODE columns
data = data_raw.copy()
data[dt_col] = data[dt_col].fillna(method='ffill')  # DT_NAME forward fill
data[feeder_col] = data[feeder_col].fillna(method='ffill')  # FEEDER_CODE forward fill

# Normalize and prepare data
data["DT_NAME"] = data[dt_col].apply(normalize)
data["FEEDER_CODE"] = data[feeder_col].astype(str).str.strip()
data["MR_NAME"] = data[mr_name_col].astype(str).str.strip()
data["MR_CODE"] = data[mr_code_col].astype(str).str.strip()
data["CONSUMERCODE"] = data[consumer_col].astype(str).str.zfill(13)
data["SP_ID"] = ""

# 🟢 Step 2: Generate unique DT_CODEs
dt_unique = data["DT_NAME"].drop_duplicates().tolist()
st.success(f"✅ {len(dt_unique)} unique DT Names. Total rows = {len(data)}")

base_code = st.text_input("Start DT_CODE (e.g., DT-101)", "DT-101")
if not re.match(r"^[A-Za-z\-]+\d+$", base_code):
    st.error("Invalid format; use DT-101 or DT100 etc.")
    st.stop()

dt_codes = increment_dt_codes(base_code, len(dt_unique))
dt_map = dict(zip(dt_unique, dt_codes))

# 🟢 Step 3: Map back DT_CODE to ALL rows
data["DT_CODE"] = data["DT_NAME"].map(dt_map)

# Inputs
subdivision = st.text_input("Subdivision")
ticket_id = st.text_input("Ticket ID")

# 🔽 Generate Excel
if st.button("🚀 Generate Excel"):
    wb = Workbook()

    # Sheets with unique DT_NAMEs
    sheet_templates = {
        "FAC_LVL_2": ("INSERT INTO CI_FAC_LVL_2 VALUES('ES','{code}',1);", ["DT_NAME", "DT_CODE", "FEEDER_CODE", "MR_NAME", "MR_CODE"]),
        "FAC_LVL_2_L": ("INSERT INTO CI_FAC_LVL_2_L VALUES('ES','{code}','ENG','{name}',1);", ["DT_NAME", "DT_CODE", "FEEDER_CODE", "MR_NAME", "MR_CODE"]),
        "FAC_LVL_1_2": ("INSERT INTO CI_FAC_LVL_1_2 VALUES('ES','{feeder}','{code}',1);", ["DT_NAME", "DT_CODE", "FEEDER_CODE", "MR_NAME", "MR_CODE"]),
        "METER_READER": ("INSERT INTO XX_METER_READER_RT VALUES ('{code}','{name}','{mrname}','{mrcode}');", ["DT_NAME", "DT_CODE", "FEEDER_CODE", "MR_NAME", "MR_CODE"]),
    }

    unique_dt = data.drop_duplicates("DT_NAME")

    for idx, (sheet_name, (template, col_list)) in enumerate(sheet_templates.items()):
        ws = wb.active if idx == 0 else wb.create_sheet(sheet_name)
        ws.title = sheet_name
        ws.append(col_list + ["QUERY"])
        for _, r in unique_dt.iterrows():
            query = template.format(
                code=r["DT_CODE"],
                feeder=r["FEEDER_CODE"],
                name=r["DT_NAME"],
                mrname=r["MR_NAME"],
                mrcode=r["MR_CODE"]
            )
            ws.append([r[c] for c in col_list] + [query])

    # ✅ SP_UPDATE – Full consumer list
    ws5 = wb.create_sheet("SP_UPDATE")
    ws5.append(["CONSUMERCODE", "DT_CODE", "FEEDER_CODE", "DT_NAME", "SP_ID", "QUERY"])
    for i, r in enumerate(data.itertuples(index=False), start=2):
        formula = (
            f'="update ci_sp set fac_lvl_1_cd = \'" & C{i} & "\', fac_lvl_2_cd= \'" & B{i} & "\' where sp_id = \'" & E{i} & "\' ;"'
        )
        ws5.append([
            r.CONSUMERCODE,
            r.DT_CODE,
            r.FEEDER_CODE,
            r.DT_NAME,
            r.SP_ID,
            formula
        ])

    # Save file
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    st.download_button(
        "⬇️ Download Excel",
        buf,
        f"DT_CREATION_{ticket_id}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
